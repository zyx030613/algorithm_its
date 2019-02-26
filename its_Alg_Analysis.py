# coding:utf-8
from __future__ import division
import logging
logger = logging.getLogger("its_Alg")
from PyQt4 import QtCore
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import numbers
import openpyxl,os,re
import openpyxl.styles as sty
#定义报告中三个sheet页的名称
DATA_STD = u'标准数据'
DATA_TEST= u'测试数据'
DATA_RESULT = u'统计结果'
#定义指标的背景颜色
TYPE_COLOR={
    #定义错误项的标记颜色
    '抓拍率':'ff0000',
    '误抓率':'808080',
    '识别率':'ffc000',
    '车型':'ffc000',
    '车标':'ffc000',
    '子品牌':'ffc000',
    '车身颜色':'ffc000',
    '接打电话':'ffc000',
    '接打电话有效率':'ffc000',
    '不系安全带':'ffc000',
    '不系安全带有效率':'ffc000',
    '主遮阳板':'ffc000',
    '副遮阳板':'ffc000',
    '纸巾盒':'ffc000',
    '挂坠':'ffc000',
    '主驾驶性别':'ffc000',
    '副驾驶性别':'ffc000'
}
#定义车牌顺序误差阈值(>=1,1为没有顺序误差)
ORDER_ERR = 4
#定义车牌相似度阈值
PLATE_SLR = 4
#定义指标字段和指标名称的对应关系
E_C = {
    'cphm':'识别率','clpp':'车标','clzpp':'子品牌','cllx':'车型',
    'csys':'车身颜色','ddh':'接打电话','aqd':'不系安全带','zzyb':'主遮阳板',
    'fzyb': '副遮阳板','zjh':'纸巾盒','gz':'挂坠','zjsxb':'主驾驶性别',
    'fjsxb':'副驾驶性别','z_cpdz':'误抓率','z_cplz':'抓拍率','ddh_yxl':'接打电话有效率',
    'aqd_yxl':'不系安全带有效率'
}
#定义excel表单字段顺序
TAB_ORDER = {
    'cphm':2,'clpp':5,'clzpp':6,'cllx':7,'csys':8,'ddh':9,'aqd':10,'zzyb':11,'fzyb':12,
    'zjh':13,'gz':14,'zjsxb':15,'fjsxb':16,'z_cpdz':2,'z_cplz':2,'pic_dir':17
}
#定义报告页面指标顺序
ORDER_REP = {
    'cphm':4,'clpp':7,'clzpp':8,'cllx':6,
    'csys':5,'ddh':9,'aqd':10,'zzyb':13,
    'fzyb': 14,'zjh':11,'gz':12,'zjsxb':15,
    'fjsxb':16,'z_cpdz':3,'z_cplz':2,'ddh_yxl':17,
    'aqd_yxl':18
}
#定义报告页面字体
FONT_TITLE = sty.Font(name='Times New Roman', bold=True, size=14)
#定义相似字符
SIM1 = ['B','8']
SIM2 = ['C','0','D','Q','U']
SIM3 = ['E','F']
SIM4 = ['G','6']
SIM5 = ['L','1','T']
SIM6 = ['H','N']
SIM7 = ['S','5']
SIM8 = ['Z','2','7']
SIMILAR = [SIM1,SIM2,SIM3,SIM4,SIM5,SIM6,SIM7,SIM8]
class Alg_Interface_Thread(QtCore.QThread):
    """docstring for BigWorkThread"""
    # 声明一个信号，同时返回一个list，同理什么都能返回啦
    finishSignal = QtCore.pyqtSignal(list)
    errSignal = QtCore.pyqtSignal(str) #返回错误信息

    # 构造函数里增加形参
    def __init__(self, excel_std,excel_test,excel_report,test_flag,rec_flag):
        '''
        
        :param excel_std: 标准数据excel路径
        :param excel_test: 测试数据excel路径
        :param excel_report: 测试报告要生成的路径
        :param test_flag: 分析模式，传1为自动分析，0为excel修改后手动分析
        :param rec_flag: 是否勾选卡口识别率，0为不勾选，否则勾选
        '''
        super(Alg_Interface_Thread, self).__init__()
        self.excel_std = excel_std
        self.excel_test = excel_test
        self.excel_report = excel_report
        #定义自动分析还是手动修改后分析，1为自动分析，0为手动修改后分析
        self.test_flag = test_flag
        self.rec_flag = rec_flag
        # self.run()

    def run(self):
        if self.test_flag:
            # 合并excel
            try:
                Do_OneExcel_Thread(self.excel_std, self.excel_test, self.excel_report)
            except:
                logger.exception('ERR_Do_OneExcel_Thread')
                self.errSignal.emit(u'合并excel失败！')
                return
            #比对分析
            try:
                if not self.rec_flag:   #如果没勾选卡口识别
                    self.analysis_res = Data_Analysis_Thread(self.excel_report)
                else:
                    self.analysis_res = Data_Analysis_REC_Thread(self.excel_report)
                    if self.analysis_res.excel_style == {}:     #如果模拟触发记录数不对，用抓拍率比对算法分信息
                        self.analysis_res = Data_Analysis_Thread(self.excel_report)
            #获取样式结果
                self.excel_style = self.analysis_res.excel_style
            except:
                logger.exception('ERR_Data_Analysis_Thread')
                self.errSignal.emit(u'数据对比分析失败！')
                return
            #样式结果写入excel
            try:
                Excel_Report_Thread(self.excel_report, self.excel_style)
            except:
                logger.exception('ERR_Excel_Report_Thread')
                self.errSignal.emit(u'样式结果写入excel失败！')
                return
        #计算各项指标
        try:
            self.analysis_res = Calc_ZB_Thread(self.excel_report)
        #获取指标样式结果
            self.excel_style = self.analysis_res.excel_style
        except:
            logger.exception('ERR_Calc_ZB_Thread')
            self.errSignal.emit(u'计算各项指标失败！')
            return
        #指标样式结果写入excel
        try:
            Excel_Report_Thread(self.excel_report, self.excel_style)
        except:
            logger.exception('ERR_Excel_Report_Thread')
            self.errSignal.emit(u'生成测试报告失败！')
            return
        self.finishSignal.emit([self.excel_report])

class Do_OneExcel_Thread():
    #将标准结果和测试结果合并，输出合并后的excel
    def __init__(self,excel_std,excel_test,excel_finish):
        # super(Do_OneExcel_Thread, self).__init__()
        self.excel_std = excel_std
        self.excel_test = excel_test
        self.excel_finish = excel_finish
        self.sheetname_finish1 = DATA_STD
        self.sheetname_finish2 = DATA_TEST
        self.start_run()
        logger.critical('Merge two excel Finish！')

    def start_run(self):
        self.l_list = [(self.excel_test,self.sheetname_finish2)
                       ,(self.excel_std,self.sheetname_finish1)]
        for paramlist in self.l_list:
            self.sheet = self._Read_Excel_Data(paramlist[0])
            self.timesort = self.time_sort(self.sheet)
            self._Write_Excel_Data(self.excel_finish,paramlist[1],self.sheet,self.timesort)

    def _Read_Excel_Data(self,excelname):
        #获取excel指定sheet页数据,返回sheet地址
        self.excelname = excelname
        self.datalist = []
        self.wb = openpyxl.load_workbook(self.excelname)
        self.sheet = self.wb.active
        return self.sheet

    def time_sort(self,sheetobj):
        #传入sheet对象，返回按时间排序后的行列表
        self.rowtimelist = []
        self.rowtimeafter = []
        self.sheetobj = sheetobj
        for rownum in range(2,self.sheetobj.max_row+1):
            self.wfdm = str(self.sheetobj.cell(row = rownum,column=4).value)
            if self.wfdm == '0': #如果违法代码是0，加入排序
                self.captime = self.sheetobj.cell(row = rownum,column=1).value
                self.rowtimelist.append((rownum,self.captime))
        self.rowtimelist.sort(key=lambda x: x[1])
        for rowtime in self.rowtimelist:
            self.rowtimeafter.append(rowtime[0])
        return self.rowtimeafter

    def _Write_Excel_Data(self,excelname,sheetname,sheetold,timesort):
        #excel指定sheet页写入指定数据
        self.excelname = excelname
        self.sheetname = sheetname
        self.sheetold = sheetold
        self.timesort = timesort
        if not os.path.exists(self.excelname):
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(self.excelname)
        self.sheetlist = self.wb.sheetnames
        if self.sheetname not in self.sheetlist:
            self.wb.create_sheet(index=0,title=self.sheetname)
        self.sheet = self.wb[self.sheetname]
        self.sheet.freeze_panes='C2'
        for i, row in enumerate(self.sheetold.iter_rows(),start=1):
            for j,cell in enumerate(row):
                if i == 1:
                    self.sheet.cell(row=1, column=j + 1, value=cell.value)
                else:
                    if i in self.timesort: #如果行号在时间排队的序列里，即卡口记录
                        self.sheet.cell(row=self.timesort.index(i)+2,column=j+1,value=unicode(cell.value).strip())
                    else:
                        pass
        self.wb.save(self.excelname)

class Data_Analysis_Thread():
    #数据比对分析，输出分析数据
    def __init__(self,excel_finish):
        # super(Data_Analysis_Thread,self).__init__()
        self.excel_finish = excel_finish
        self.sheet_std_name = DATA_STD
        self.sheet_test_name = DATA_TEST
        self.excel_style = {}
        self.start_run()
        logger.critical('Data Analysis Finish!')

    def start_run(self):
        self.platelist_std,self.sheet_std = self.Read_Excel_Platelist(self.excel_finish,self.sheet_std_name)
        self.platelist_test,self.sheet_test = self.Read_Excel_Platelist(self.excel_finish, self.sheet_test_name)
        self.Cap_Rec_Rate(self.platelist_std,self.platelist_test)

    def Read_Excel_Platelist(self,excelname,sheetname):
        #传入excel和sheet页名称,输出车牌列表
        self.excelname = excelname
        self.sheetname = sheetname
        self.wb = openpyxl.load_workbook(self.excelname)
        self.sheet = self.wb[self.sheetname]
        self.platelist = [x.value for x in self.sheet[get_column_letter(TAB_ORDER['cphm'])] if x.value][1:]
        return self.platelist,self.sheet

    def Cap_Rec_Rate(self,list_std, list_test):
        #传入标准数据和测试数据的车牌列表
        self.sheet_std_typelist = []
        self.sheet_test_typelist = []
        self.pair_first = self.first_filter(list_std, list_test)
        for zbname in E_C.keys():
            if zbname not in ['z_cpdz','z_cplz','aqd_yxl','ddh_yxl']:
                self.excel_cell_compar(self.pair_first, zbname)
        #将三个sheet页要修改的风格放入字典
        self.excel_style[DATA_STD] = self.sheet_std_typelist
        self.excel_style[DATA_TEST] = self.sheet_test_typelist

    def excel_cell_compar(self,pairlist,zbname):
        #比对两个sheet页面同一条记录的指定字段是否相同
        #传入对比行的列表和指标名称
        self.pairlist = pairlist
        self.zbname = zbname
        self.col_num = TAB_ORDER[self.zbname]
        for row_tuple in self.pairlist:
            std_value = self.sheet_std.cell(row=row_tuple[0]+2,column=self.col_num).value#车牌index是从0开始，实际车牌从第2行开始
            test_value = self.sheet_test.cell(row=row_tuple[1]+2,column=self.col_num).value
            if self.zbname == 'cphm':
                if std_value != test_value:#如果是车牌且比对结果不一样
                    self.test_rec_err = [row_tuple[1]+2,self.col_num,'',TYPE_COLOR[E_C[self.zbname]]]
                    self.sheet_test_typelist.append(self.test_rec_err)
            else:
                if test_value not in std_value:#如果不是车牌且标准结果不包含识别结果
                    self.test_rec_err = [row_tuple[1]+2,self.col_num,'',TYPE_COLOR[E_C[self.zbname]]]
                    self.sheet_test_typelist.append(self.test_rec_err)

    def plate_similar_1(self,plate1, plate2):
        # 字母加数字部分，相同位置字符一样或相似的数量大于等于n个，代表相似,返回真，否则返回假
        plate1_s = re.findall(r'[A-Z0-9]', plate1)
        plate2_s = re.findall(r'[A-Z0-9]', plate2)
        if len(plate1_s)==0 or len(plate2_s)==0:#如果车牌中无数字或字母，不进行相似匹配
            return False
        similar_s = [plate1_s[i] for i in range(min(len(plate1_s), len(plate2_s)))
                     if self._similar_special(plate1_s[i],plate2_s[i])]
        if len(similar_s) >= PLATE_SLR:
            return True
        else:
            return False

    def plate_similar_2(self,plate1, plate2):
        # 字母加数字部分，存在相同的连续字符n个，代表相似,返回真，否则返回假
        plate1_s = re.findall(r'[A-Z0-9]+', plate1)
        plate2_s = re.findall(r'[A-Z0-9]+', plate2)
        if len(plate1_s)==0 or len(plate2_s)==0:#如果车牌中无数字或字母，不进行相似匹配
            return False
        plate1_s = plate1_s[0]
        plate2_s = plate2_s[0]
        l_plate1 = len(plate1_s)
        for i in range(l_plate1):
            if i + PLATE_SLR <= l_plate1:
                sub_plate1 = plate1_s[i:i + PLATE_SLR]
                if sub_plate1 in plate2_s:
                    return True
            else:
                return False

    def _similar_special(self,char1,char2):
        #传入两个字符，一样或者相似返回真，否则返回假
        self.char1 = char1
        self.char2 = char2
        if self.char1 == self.char2:
            return True
        else:
            self.simflag = 0
            for simlist in SIMILAR:
                if self.char1 in simlist and self.char2 in simlist:
                    self.simflag = 1
                    return True
                else:
                    pass
            if not self.simflag:
                return False

    def first_filter(self,list_std, list_test):
        # 传入标准列表和待测列表
        # 输出确认抓拍列表，确认识别正确列表，识别错误列表，多抓列表
        self.len_std = len(list_std)
        self.len_test = len(list_test)
        self.pair_first = []
        self.list_waittest = {key: value for key, value in enumerate(list_test)}
        self.list_matched_index_test = []
        #定义test中匹配上的车牌index
        self.find_index_in_test = 0
        for self.plate_std_index, plate_num in enumerate(list_std):
            # 设置找到标记=0为没有找到,1为找到
            self.find_flag = 0
            #在上一个已匹配的车牌位置上下浮动ORDER_ERR范围内查找下一个车牌,已匹配过的不能重复匹配
            self.find_start_index = \
                self.find_index_in_test-ORDER_ERR if self.find_index_in_test-ORDER_ERR >=0 else 0
            self.find_end_index = \
                self.find_index_in_test+ORDER_ERR+1 if self.find_index_in_test+ORDER_ERR+1 <=self.len_test else self.len_test
            logger.debug(plate_num+' will First_find in'+str([self.find_index_in_test,self.find_start_index,self.find_end_index]))
            logger.debug( 'Have matched list:'+str(self.list_matched_index_test))
            self.thisfind_list = [(x,list_test[x]) for x in range(self.find_start_index,self.find_end_index) if
                                      x not in self.list_matched_index_test]
            logger.debug('this time need be matched plate:'+str(self.thisfind_list))
            for (plate_index,plate_value) in self.thisfind_list:
                if plate_num == plate_value:
                    #如果在范围内有相同车牌，放入抓拍正确列表和识别正确列表
                    self.find_flag = 1
                    #已匹配车牌在test中的index
                    self.find_index_in_test = plate_index
                    logger.debug(plate_num+'......First Compar and Rec OK')
                    break
            else:
                #判断查找范围内是否存在相似车牌
                for (self.plate_test_index,plate_num_test) in self.thisfind_list:
                    #相似车牌两种判断逻辑，一种成功就算相似
                    self.rec_err_result = self.plate_similar_1(plate_num,plate_num_test)\
                                     or self.plate_similar_2(plate_num,plate_num_test)
                    if self.rec_err_result:
                        # 如果存在相似车牌
                        # self.index_end_in_std = self.plate_std_index+ORDER_ERR+1\
                        #     if self.plate_std_index+ORDER_ERR+1<=self.len_std else self.len_std
                        if plate_num_test not in list_std[self.plate_std_index:]:
                            #且此相似车牌在标准表中要匹配车牌的后ORDER_ERR中不存在
                            self.find_flag = 1
                            self.find_index_in_test = self.plate_test_index
                            logger.debug(plate_num + '......First Compar but Rec ERR '+ plate_num_test+'!!!!!')
                            break
                        else:
                            pass
                    else:
                        #如果车牌不相似，继续比较下一个
                        pass
                if not self.find_flag and plate_num!=u'未知':
                    #非未知车牌的二次匹配：如果范围内无相似车牌，再在匹配上的车牌往后整个列表中查找有无一样，如果有，再判断std中下一个车牌
                    #在test中找到的车牌周边是否存在一样
                    logger.debug(plate_num + ' First_find FAIL!,will Second_find in '
                                  + str([self.find_index_in_test,self.len_test]))
                    for self.plate_test_index,plate_num_test in enumerate(list_test[self.find_end_index:],start=self.find_index_in_test):
                        if plate_num_test == plate_num:
                            self.find_index_in_test_start = self.plate_test_index-ORDER_ERR \
                                if self.plate_test_index-ORDER_ERR>=0 else 0
                            self.find_index_in_test_end = self.plate_test_index + ORDER_ERR +1 \
                                if self.plate_test_index + ORDER_ERR+1 <= self.len_test else self.len_test
                            self.next_plate = self.plate_std_index + 1 if self.plate_std_index +1 < self.len_std else self.plate_std_index - 1
                            if list_std[self.next_plate] in list_test[self.find_index_in_test_start:self.find_index_in_test_end]:
                                self.find_flag = 1
                                self.find_index_in_test = self.plate_test_index
                                logger.debug(plate_num + '......Second Compar and Rec OK')
                                break
                        else:
                            #继续匹配下一个车牌
                            pass
                #如果范围内无相似车牌，放入漏抓列表
                if not self.find_flag:
                    self.celltype = [self.plate_std_index + 2, 2, '', TYPE_COLOR['抓拍率']]
                    self.sheet_std_typelist.append(self.celltype)
                    logger.debug(plate_num + ' Second Compar FAIL,miss Capture!!!')
            if self.find_flag:
                self.pair_first.append((self.plate_std_index, self.find_index_in_test))
                logger.debug('delete matched index in test:'+ str(self.find_index_in_test))
                del self.list_waittest[self.find_index_in_test]
                self.list_matched_index_test.append(self.find_index_in_test)
        for capmore_index in self.list_waittest.keys():
            # 将多抓车牌标注颜色
            self.celltype = [capmore_index + 2, 2, '', TYPE_COLOR['误抓率']]
            self.sheet_test_typelist.append(self.celltype)
        return self.pair_first

class Excel_Report_Thread():
    #根据输出结果在excel中标注颜色
    def __init__(self,excelname,dict_excel_style):
        # super(Excel_Report_Thread,self).__init__()
        self.excelname = excelname
        self.excel_style = dict_excel_style
        self.start_run()
        logger.critical('Excel Style Finish!')

    def start_run(self):
        self.Write_Excel_Style(self.excelname,self.excel_style)

    def Write_Excel_Style(self,excelname,dict_excel_style):
        #根据传入的excel格式字典，对excel进行标记
        self.excelname = excelname
        self.excel_style = dict_excel_style
        self.wb = openpyxl.load_workbook(self.excelname)
        self.sheets = self.wb.sheetnames
        for sheetname in self.excel_style.keys():
            if sheetname not in self.sheets:
                self.wb.create_sheet(index=0, title=sheetname)
            if sheetname in self.sheets and sheetname == DATA_RESULT:
                #如果操作的是报告页面，且报告页面已经存在，先删除再重新生成
                self.wb.remove(self.wb[sheetname])
                self.wb.create_sheet(index=0, title=sheetname)
            self.sheet = self.wb[sheetname]
            for cellstyle in self.excel_style[sheetname]:
                if cellstyle[3] == 'ffffff':
                    self.sheet.cell(row=cellstyle[0], column=cellstyle[1]).fill = sty.PatternFill(fill_type=None,
                                                                                                  fgColor=cellstyle[3])
                else:
                    self.sheet.cell(row=cellstyle[0], column=cellstyle[1]).fill = sty.PatternFill(fill_type='solid',
                                                                                              fgColor=cellstyle[3])
                if cellstyle[2]!='':
                    self.sheet.cell(row=cellstyle[0], column=cellstyle[1]).font = FONT_TITLE
                    self.sheet.cell(row=cellstyle[0], column=cellstyle[1], value=cellstyle[2])
            if sheetname == DATA_RESULT:#如果是报告页面
                self.sheet.column_dimensions['A'].width = 25    #设置A列宽度
                self.sheet.column_dimensions['B'].width = 12    #设置B列宽度
                self.sheet.column_dimensions['C'].width = 20  # 设置B列宽度
                for row in range(self.sheet.max_row):
                    self.sheet.cell(row=row+1,column=2).number_format = numbers.FORMAT_PERCENTAGE_00
                self.sheet = self.Do_Excel_Chart(self.sheet,2,2,self.sheet.max_row,2)
        self.wb.save(self.excelname)

    def Do_Excel_Chart(self,sheet,row_start,col_start,row_end,col_end):
        #将数据绘制成图表
        self.sheet = sheet
        self.row_start = row_start
        self.col_start = col_start
        self.row_end = row_end
        self.col_end = col_end
        self.refObj = Reference(self.sheet,self.col_start,self.row_start,self.col_end,self.row_end)
        self.yObj = Reference(self.sheet, min_col=1, min_row=self.row_start, max_row=self.row_end)
        self.seriesObj = Series(self.refObj,title='TD')
        self.chartObj = BarChart()
        self.chartObj.title = u'指标统计结果'
        self.chartObj.append(self.seriesObj)
        self.chartObj.y_axis.title = u'指标结果'
        self.chartObj.x_axis.title = u'指标类别'
        self.chartObj.set_categories(self.yObj)
        self.sheet.add_chart(self.chartObj,'E2')
        return self.sheet

class Calc_ZB_Thread():
    #根据颜色标注结果，并生成指标统计结果及报告页面样式
    def __init__(self,excelname):
        # super(Calc_ZB_Thread,self).__init__()
        self.excelname = excelname
        self.excel_style = {}
        self.start_run()
        logger.critical('Calc Result Finish!')

    def start_run(self):
        self.zb_type_list = [[1,1,u'指标类别','ffffff'],[1,2,u'指标结果','ffffff'],[1,3,u'计算方式','ffffff']]
        self.wb = openpyxl.load_workbook(self.excelname)
        self.sheet_std = self.wb[DATA_STD]
        self.sheet_test = self.wb[DATA_TEST]
        ####开始计算安全带和打电话的有效率##########
        self.ddh_capnum = 0 #抓拍到的打电话和不系安全带数量
        self.aqd_capnum = 0
        self.ddh_yxnum = 0 #打电话和安全带有效的数量
        self.aqd_yxnum = 0
        for rownum in range(1,self.sheet_test.max_row):
            r_ddh = self.sheet_test.cell(row=rownum+1,column = TAB_ORDER['ddh']).value
            r_aqd = self.sheet_test.cell(row=rownum+1,column = TAB_ORDER['aqd']).value
            if self.cell_color(self.sheet_test,rownum+1,TAB_ORDER['cphm'])[2:].upper()!=TYPE_COLOR['误抓率'].upper():
            # 如果不是误抓车牌
                if r_ddh == u'是':
                    self.ddh_capnum = self.ddh_capnum + 1
                    if self.cell_color(self.sheet_test,rownum+1,TAB_ORDER['ddh'])=='00000000':
                        self.ddh_yxnum = self.ddh_yxnum + 1
                if r_aqd == u'否':
                    self.aqd_capnum = self.aqd_capnum + 1
                    if self.cell_color(self.sheet_test,rownum+1,TAB_ORDER['aqd'])=='00000000':
                        self.aqd_yxnum = self.aqd_yxnum + 1
            else:
                pass
        self.ddh_yxl = self.ddh_yxnum / self.ddh_capnum if self.ddh_capnum != 0 else 0
        self.aqd_yxl = self.aqd_yxnum / self.aqd_capnum if self.aqd_capnum != 0 else 0
        for zbname in ['aqd_yxl','ddh_yxl']:
            self.zb_type = [ORDER_REP[zbname], 1, E_C[zbname], TYPE_COLOR[E_C[zbname]]]
            self.zb_type_list.append(self.zb_type)
            self.zb_type = [ORDER_REP[zbname], 2, self.ddh_yxl, 'ffffff']
            self.zb_type_list.append(self.zb_type)
            if zbname == 'aqd_yxl':
                self.zb_type = [ORDER_REP[zbname], 3, str(self.aqd_yxnum) + '/' + str(self.aqd_capnum),
                            'ffffff']  # 将多抓计算方式写入表格
            elif zbname == 'ddh_yxl':
                self.zb_type = [ORDER_REP[zbname], 3, str(self.ddh_yxnum) + '/' + str(self.ddh_capnum),
                                'ffffff']  # 将多抓计算方式写入表格
            else:
                self.zb_type = []
            self.zb_type_list.append(self.zb_type)
        ####结束计算安全带和打电话的有效率#########################
        self.count_std = self.sheet_std.max_row - 1#标准过车数
        self.count_test = self.sheet_test.max_row - 1#实际抓拍数
        self.count_more = self.count_color(self.sheet_test, 'z_cpdz')#多抓数
        #多抓率
        self.rate_more = self.count_more / self.count_test if self.count_test!=0 else 0
        self.zb_type = [ORDER_REP['z_cpdz'],1,E_C['z_cpdz'],TYPE_COLOR[E_C['z_cpdz']]]
        self.zb_type_list.append(self.zb_type)
        self.zb_type = [ORDER_REP['z_cpdz'],2,self.rate_more,'ffffff']
        self.zb_type_list.append(self.zb_type)
        self.zb_type = [ORDER_REP['z_cpdz'],3,str(self.count_more)+ '/'+ str(self.count_test),'ffffff'] #将多抓计算方式写入表格
        self.zb_type_list.append(self.zb_type)
        self.count_real = self.count_test - self.count_more#实际有效抓拍数
        self.type_row_num = 2
        for zdname in ORDER_REP.keys():
            if zdname not in ['z_cpdz','ddh_yxl','aqd_yxl']: #如果不是抓拍率、打电话有效率、安全带有效率
                if zdname == 'z_cplz':
                    self.count = self.count_color(self.sheet_std,zdname)
                    #抓拍率
                    self.rate =  (self.count_std - self.count)/self.count_std
                    self.zb_type_res = [ORDER_REP[zdname], 3, '('+str(self.count_std)+ '-' +str(self.count)+')/'+str(self.count_std), 'ffffff']
                    self.zb_type_list.append(self.zb_type_res)
                else:
                    self.count = self.count_color(self.sheet_test,zdname)
                    #车牌及其他项识别率
                    self.rate = (self.count_real - self.count) / self.count_real if self.count_real!=0 else 0
                    self.zb_type_res = [ORDER_REP[zdname], 3,'('+str(self.count_real)+ '-' +str(self.count)+')/'+str(self.count_real), 'ffffff']
                    self.zb_type_list.append(self.zb_type_res)
                self.zb_type_name = [ORDER_REP[zdname], 1, E_C[zdname],TYPE_COLOR[E_C[zdname]]]
                self.zb_type_res = [ORDER_REP[zdname], 2, self.rate, 'ffffff']
                self.zb_type_list.append(self.zb_type_name)
                self.zb_type_list.append(self.zb_type_res)
                self.type_row_num = self.type_row_num + 1
        self.excel_style[DATA_RESULT]=self.zb_type_list

    def count_color(self,sheet,zbname):
        #输入sheet对象和列名称，返回错误个数
        self.sheet = sheet
        self.zbname = zbname
        self.zb_col = TAB_ORDER[self.zbname]
        self.color_code = TYPE_COLOR[E_C[self.zbname]]
        self.max_row = self.sheet.max_row
        self.count = 0
        for rownum in range(2,self.max_row+1):
            self.c_cell = self.sheet.cell(row=rownum,column=self.zb_col)
            self.fill = self.c_cell.fill
            self.c_cell_bg = self.fill.start_color.rgb
            if self.c_cell_bg[2:].upper() == self.color_code.upper():
                self.count = self.count + 1
        return self.count
    def cell_color(self,sheet,rownum,colnum):
        #返回指定单元格的颜色代码
        self.sheet = sheet
        self.c_cell = self.sheet.cell(row=rownum,column=colnum)
        self.fill = self.c_cell.fill
        self.c_cell_bg = self.fill.start_color.rgb
        return self.c_cell_bg

class Data_Analysis_REC_Thread():
    #勾选卡口识别率时，使用此分析算法
    def __init__(self,excel_finish):
        # super(Data_Analysis_Thread,self).__init__()
        self.excel_finish = excel_finish
        self.sheet_std_name = DATA_STD
        self.sheet_test_name = DATA_TEST
        self.excel_style = {}
        self.finish_flag = 1
        self.start_run()
        logger.critical('Data Analysis Finish!')

    def start_run(self):
        # self.sheet_std_typelist = []
        self.sheet_test_typelist = []
        self.wb = openpyxl.load_workbook(self.excel_finish)
        self.sheet_std = self.wb[self.sheet_std_name]
        self.sheet_test = self.wb[self.sheet_test_name]
        if self.sheet_std.max_row != self.sheet_test.max_row:
            logger.error('capnum error in testdata!')
            return
        else:
            for zbname in E_C.keys():
                if zbname not in ['z_cpdz', 'z_cplz', 'aqd_yxl', 'ddh_yxl']:
                    self.excel_cell_compar_REC(zbname)
        self.excel_style[DATA_TEST] = self.sheet_test_typelist

    def excel_cell_compar_REC(self,zbname):
        #比对两个sheet页面同一条记录的指定字段是否相同
        #传入对比行的列表和指标名称
        self.zbname = zbname
        self.col_num = TAB_ORDER[self.zbname]
        for row in range(1, self.sheet_std.max_row):
            std_value = self.sheet_std.cell(row=row + 1, column=self.col_num).value  # 车牌index是从0开始，实际车牌从第2行开始
            test_value = self.sheet_test.cell(row=row + 1, column=self.col_num).value
            if self.zbname == 'cphm':
                if std_value != test_value:  # 如果是车牌且比对结果不一样
                    self.test_rec_err = [row + 1, self.col_num, '', TYPE_COLOR[E_C[self.zbname]]]
                    self.sheet_test_typelist.append(self.test_rec_err)
            else:
                if test_value not in std_value:  # 如果不是车牌且标准结果不包含识别结果
                    self.test_rec_err = [row + 1, self.col_num, '', TYPE_COLOR[E_C[self.zbname]]]
                    self.sheet_test_typelist.append(self.test_rec_err)

if __name__ == '__main__':
    excel1 = r'C:\Users\wangwei\Desktop\autotest\excel_std.xlsx'
    excel2 = r'C:\Users\wangwei\Desktop\autotest\excel_test.xlsx'
    excel3 = r'C:\Users\wangwei\Desktop\autotest\alg_report.xlsx'
    Alg_Interface_Thread(excel1,excel2,excel3,1,0)

