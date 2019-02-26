# -*- coding: utf-8 -*-
import os
import errno
import logging
from openpyxl import Workbook, load_workbook


logger = logging.getLogger('Excelora')

class XLS_WRITE:
	def __init__(self):
		if os.name is "nt":
			self.tmpDir = "Temp"
		else:
			self.tmpDir = "tmp"

	def open_xlsx(self,path_dst,):
		# host=result_list[0]
		host = 'sheet'
		self.path_dst = path_dst
		path_all = os.path.split(path_dst)
		if not os.path.exists(path_all[0]):
			try:
				os.mkdir(path_all[0])
			except:
				logger.debug('mkdir failed!')
		else:
			pass

		if not os.path.exists(path_dst):
			try:
				self.wb = Workbook()
				try:
					self.w_xls = self.wb.active  # create_sheet(title=host)
					self.rows = 1
				except Exception as aaa:
					logger.debug('ex open excel :%s' % aaa)
					self.w_xls = self.wb.active  # get_sheet_by_name(host)
					self.rows = self.w_xls.max_row
			except OSError as exc:
				if exc.errno == errno.EEXIST and os.path.isdir(path_dst):
					self.wb = load_workbook(path_dst)
					try:
						self.w_xls = self.wb.active  # create_sheet(title=host)
						self.rows = 1
					except:
						self.w_xls = self.wb.active  # get_sheet_by_name(host)
						self.rows = self.w_xls.max_row
					pass
				else:
					raise
		else:
			self.wb = load_workbook(path_dst)
			try:
				self.w_xls = self.wb.active  # create_sheet(title=host)
				self.rows = self.w_xls.max_row
			except:
				self.w_xls = self.wb.active  # get_sheet_by_name(host)
				self.rows = self.w_xls.max_row
			# w_xls = wb.active#get_sheet_by_name(host)

	def read_its_xlsx(self):
		a1 = self.w_xls.cell(1,1)
		print a1.value,type(a1.value)
		return a1.value



	def wirte_its_in(self, result_list):
		"""
		执行excel结果写入
		Excel路径： path_dst:
		socket结果： result_list:
		无返回值:
		"""
		self.rows = self.w_xls.max_row
		result_dict = result_list[1]
		w_xls =self.w_xls
		### 写入第一行内容
		self.w_xls.cell(1, 1, u'抓拍时间')
		self.w_xls.cell(1, 2, u'车牌号码')
		self.w_xls.cell(1, 3, u'车道名称')
		self.w_xls.cell(1, 4, u'违法代码')
		self.w_xls.cell(1, 5, u'车辆品牌')
		self.w_xls.cell(1, 6, u'车辆子品牌')
		self.w_xls.cell(1, 7, u'车辆类型')
		self.w_xls.cell(1, 8, u'车身颜色')
		self.w_xls.cell(1, 9, u'是否打电话')
		self.w_xls.cell(1, 10, u'主驾驶安全带')
		self.w_xls.cell(1, 11, u'主驾驶遮阳板')
		self.w_xls.cell(1, 12, u'副驾驶遮阳板')
		self.w_xls.cell(1, 13, u'纸巾盒')
		self.w_xls.cell(1, 14, u'挂坠')
		self.w_xls.cell(1, 15, u'主驾驶性别')
		self.w_xls.cell(1, 16, u'副驾驶性别')
		self.w_xls.cell(1, 17, u'图片路径')
		# w_xls.col(0).width = 10000

		rows = self.rows+1
		w_xls.cell(rows, 1, result_dict['snap_time'])
		if result_dict['cplate']!= 'np':
			w_xls.cell(rows, 2, result_dict['cplate'].decode('gbk'))
		else:
			w_xls.cell(rows, 2, '未知')
		w_xls.cell(rows, 3, result_dict['road_num'])
		w_xls.cell(rows, 4, result_dict['alarm_code'])
		w_xls.cell(rows, 5, result_dict['car_brand'])
		w_xls.cell(rows, 6, result_dict['car_brand_pro'])
		w_xls.cell(rows, 7, result_dict['car_type'])
		w_xls.cell(rows, 8, result_dict['car_color'])
		w_xls.cell(rows, 9, result_dict['do_phone'])
		w_xls.cell(rows, 10, result_dict['main_safe_belt'])
		w_xls.cell(rows, 11, result_dict['main_sun'])
		w_xls.cell(rows, 12, result_dict['sub_sun'])
		w_xls.cell(rows, 13, result_dict['paper_box'])
		w_xls.cell(rows, 14, result_dict['pendant'])
		w_xls.cell(rows, 15, result_dict['main_sex'])
		w_xls.cell(rows, 16, result_dict['sub_sex'])
		w_xls.cell(rows, 17, result_dict['pic_path'].decode('gbk'))
		# Save the file
		# wb.save(path_dst)
		logger.info("Excel write over, Did't save!")

	def save_xlsx(self):
		try:
			self.wb.save(self.path_dst)
			return 1
		except Exception as eee:
			logger.error('Err with save ora xlsx: %s' %eee)
			return 0


#x= XLS_WRITE()
#x.wirte_its_in('F:\\result_z.xlsx',['F:\\result_z.xlsx',[1,2,3]])